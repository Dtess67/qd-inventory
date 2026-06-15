#!/usr/bin/env python3
"""Regenerate the diffable / recoverable mirrors from the master workbook.
Run after every inventory edit:  python scripts/export_mirrors.py
Master of record: data/QD_Electronics_Inventory.xlsx
Outputs (git-tracked, text, diffable): QD_Inventory.csv, QD_Flags.csv, QD_Inventory_SNAPSHOT.md
"""
import csv, os
from openpyxl import load_workbook
from datetime import date

HERE=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA=os.path.join(HERE,"data")
SRC=os.path.join(DATA,"QD_Electronics_Inventory.xlsx")

wb=load_workbook(SRC)
inv=wb["Inventory"]; fl=wb["Flags"]; rm=wb["Read Me"]

lastupd=""
for row_ in rm.iter_rows():
    for cell in row_:
        if cell.value=="Last updated":
            lastupd=str(rm.cell(row=cell.row,column=2).value or "")

def dump_csv(ws, path, header_row, ncols):
    with open(path,"w",newline="") as f:
        w=csv.writer(f)
        for r in range(header_row, ws.max_row+1):
            vals=[ws.cell(row=r,column=c).value for c in range(1,ncols+1)]
            if any(v not in (None,"") for v in vals):
                w.writerow(["" if v is None else v for v in vals])
dump_csv(inv, os.path.join(DATA,"QD_Inventory.csv"), 4, 13)
dump_csv(fl,  os.path.join(DATA,"QD_Flags.csv"), 4, 7)

inv_items=[r for r in range(5,inv.max_row+1) if inv.cell(row=r,column=2).value]
fl_items=[r for r in range(5,fl.max_row+1) if fl.cell(row=r,column=1).value]
md=[f"# QD Electronics Inventory — text snapshot ({date.today().isoformat()})","",
    "Plain-text fallback of the master spreadsheet. If the .xlsx ever won't open, this and the CSVs hold the same data.","",
    f"**Totals:** {len(inv_items)} inventory line items · {len(fl_items)} flags/questions","",
    f"**Latest summary:** {lastupd}","","## Inventory",""]
for r in inv_items:
    g=lambda c: inv.cell(row=r,column=c).value
    md+=[f"### {g(1)}. {g(2)}  —  [{g(13)}]",
         f"- **Qty:** {g(5)} {g(6) or ''} · **Subsystem:** {g(4)} · **Mfr/Part:** {g(3)}",
         f"- **QD verdict:** {g(8)}",
         f"- **Notes:** {g(9)}",""]
md+=["## Flags / Design Roadmap",""]
for r in fl_items:
    g=lambda c: fl.cell(row=r,column=c).value
    md+=[f"### {g(1)}  ({g(2)} · {g(3)} · {g(4)})", f"- **Flag:** {g(5)}"]
    if g(6): md.append(f"- **Related:** {g(6)}")
    if g(7): md.append(f"- **Action:** {g(7)}")
    md.append("")
with open(os.path.join(DATA,"QD_Inventory_SNAPSHOT.md"),"w") as f:
    f.write("\n".join(md))
print(f"mirrors regenerated: {len(inv_items)} items, {len(fl_items)} flags")
